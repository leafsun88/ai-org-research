#!/usr/bin/env python3
"""Generate Excel: 100家港美股投研标的 (Meta~39, Shopify~38, NVIDIA~23, 有重叠)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ===== Styles =====
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="Arial", size=10)
META_FILL = PatternFill("solid", fgColor="FCE4D6")      # Orange - Meta类
SHOP_FILL = PatternFill("solid", fgColor="E2EFDA")      # Green - Shopify类
NV_FILL = PatternFill("solid", fgColor="D6DCE4")        # Gray-blue - NVIDIA类
MULTI_FILL = PatternFill("solid", fgColor="D6E4F0")     # Light blue - 多类
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

# ===== 100家公司数据 =====
# (name, ticker, exchange, mcap, category, sector, note)
# category: M=Meta类, S=Shopify类, N=NVIDIA类

DATA = [
    # ── Meta类 (39家) ──
    ("Alphabet (Google/YouTube)", "GOOGL", "NASDAQ", "$3.5T", "MN", "大型科技平台", "广告+云+AI芯片(TPU)三位一体"),
    ("Microsoft", "MSFT", "NASDAQ", "$3T", "MN", "大型科技平台", "LinkedIn+Azure AI云+Maia AI芯片"),
    ("Amazon", "AMZN", "NASDAQ", "$2T", "MN", "电商/云/广告", "AWS云+$56B广告+Trainium AI芯片"),
    ("Tencent (腾讯)", "0700.HK", "HKEX", "$500B", "M", "大型科技平台", "WeChat=中国Meta，营销服务广告超千亿"),
    ("Netflix", "NFLX", "NASDAQ", "$400B", "M", "流媒体", "流媒体AVOD广告+全球最大内容飞轮"),
    ("Alibaba (阿里巴巴)", "BABA", "NYSE", "$300B", "MS", "电商/云", "阿里妈妈最大电商广告"),
    ("Salesforce", "CRM", "NYSE", "$250B", "MS", "CRM/SaaS", "企业CRM平台+AppExchange生态"),
    ("Palantir Technologies", "PLTR", "NASDAQ", "$200B", "MN", "AI平台", "AI平台(AIP+Foundry)"),
    ("Adobe", "ADBE", "NASDAQ", "$160B", "M", "创意/营销云", "Experience Cloud广告分析"),
    ("PDD Holdings (拼多多)", "PDD", "NASDAQ", "$140B", "M", "电商/社交电商", "拼多多社交电商，赛马机制组织"),
    ("AppLovin", "APP", "NASDAQ", "$130B", "M", "AdTech", "AXON AI广告引擎"),
    ("Meituan (美团)", "3690.HK", "HKEX", "$100B", "M", "本地生活", "中国本地生活超级平台"),
    ("MercadoLibre", "MELI", "NASDAQ", "$100B", "MS", "电商/FinTech", "拉美电商+支付+广告+物流超级平台"),
    ("Spotify", "SPOT", "NYSE", "$100B", "M", "流媒体/音频", "音频流媒体+AI广告+播客生态"),
    ("Snowflake", "SNOW", "NYSE", "$60B", "MSN", "数据云", "云数据平台"),
    ("Sea Limited", "SE", "NYSE", "$50B", "MS", "超级平台", "Shopee电商+Garena游戏+SeaMoney"),
    ("Kuaishou (快手)", "1024.HK", "HKEX", "$50B", "M", "短视频", "中国短视频第二，AI推荐驱动广告"),
    ("JD.com (京东)", "JD", "NASDAQ", "$50B", "M", "电商", "中国直销电商广告"),
    ("Take-Two Interactive", "TTWO", "NASDAQ", "$30B", "M", "游戏", "含Zynga移动游戏广告"),
    ("Baidu (百度)", "BIDU", "NASDAQ", "$30B", "M", "搜索/AI", "中国最大搜索广告"),
    ("NetEase (网易)", "NTES", "NASDAQ", "$30B", "M", "游戏/内容", "中国游戏+内容平台"),
    ("Trip.com", "TCOM", "NASDAQ", "$30B", "M", "旅游", "中国最大旅游平台"),
    ("Warner Bros. Discovery", "WBD", "NASDAQ", "$20B", "M", "流媒体/内容", "Max流媒体广告层"),
    ("Roblox", "RBLX", "NYSE", "$20B", "M", "游戏/元宇宙", "UGC虚拟世界平台"),
    ("HubSpot", "HUBS", "NYSE", "$20B", "MS", "CRM/营销SaaS", "入站营销+CRM"),
    ("The Trade Desk", "TTD", "NASDAQ", "$18B", "M", "AdTech/DSP", "最大独立DSP"),
    ("Snap", "SNAP", "NYSE", "$18B", "M", "社交媒体", "Snapchat年轻用户+AR+广告变现"),
    ("Reddit", "RDDT", "NYSE", "$18B", "M", "社交/社区", "社区聚合+广告高速增长"),
    ("Grab Holdings", "GRAB", "NASDAQ", "$15B", "MS", "超级App", "东南亚超级App，出行/外卖/金融"),
    ("Pinterest", "PINS", "NYSE", "$15B", "M", "社交/发现", "视觉发现平台，购买意图数据强"),
    ("Tencent Music (腾讯音乐)", "TME", "NYSE", "$10B", "M", "音乐/社交", "腾讯生态音乐+社交娱乐"),
    ("Roku", "ROKU", "NASDAQ", "$9B", "M", "CTV", "CTV操作系统+广告平台"),
    ("Klaviyo", "KVYO", "NYSE", "$9B", "MS", "电商营销", "电商邮件+SMS营销"),
    ("Maplebear (Instacart)", "CART", "NASDAQ", "$8B", "M", "零售媒体", "零售媒体+购物数据广告"),
    ("Paramount Global", "PARA", "NASDAQ", "$8B", "M", "流媒体/内容", "Paramount+广告支持层"),
    ("Bilibili", "BILI", "NASDAQ", "$8B", "M", "社交/视频", "中国Z世代视频社区"),
    ("Unity Software", "U", "NYSE", "$7B", "M", "游戏引擎/AdTech", "游戏引擎+IronSource广告"),
    ("Match Group", "MTCH", "NASDAQ", "$6B", "M", "约会", "Tinder/Hinge矩阵"),
    ("Kanzhun (Boss直聘)", "BZ", "NASDAQ", "$5B", "M", "招聘", "中国招聘双边平台，AI匹配"),

    # ── Shopify类 (38家) ──
    ("Oracle", "ORCL", "NYSE", "$350B", "SN", "ERP/云", "ERP+NetSuite+OCI AI"),
    ("SAP SE", "SAP", "NYSE", "$250B", "S", "ERP", "全球最大ERP"),
    ("ServiceNow", "NOW", "NYSE", "$200B", "SN", "企业工作流", "企业工作流+Now Assist AI"),
    ("Intuit", "INTU", "NASDAQ", "$190B", "S", "金融SaaS", "QuickBooks+TurboTax"),
    ("Booking Holdings", "BKNG", "NASDAQ", "$150B", "S", "旅游平台", "全球最大旅行平台"),
    ("PayPal", "PYPL", "NASDAQ", "$70B", "S", "支付", "商家支付网关+BNPL"),
    ("Fiserv", "FI", "NASDAQ", "$70B", "S", "支付", "Clover POS+银行核心系统"),
    ("Airbnb", "ABNB", "NASDAQ", "$70B", "S", "住宿平台", "住宿版Shopify"),
    ("Workday", "WDAY", "NASDAQ", "$60B", "SN", "HR/财务SaaS", "人力+财务操作系统"),
    ("Cloudflare", "NET", "NYSE", "$50B", "S", "CDN/边缘云", "网络+安全+边缘计算"),
    ("Datadog", "DDOG", "NASDAQ", "$40B", "SN", "可观测性", "云监控+AI可观测性"),
    ("Veeva Systems", "VEEV", "NYSE", "$35B", "S", "生命科学SaaS", "生命科学垂直SaaS"),
    ("Block (Square)", "SQ", "NYSE", "$35B", "S", "支付/SaaS", "中小商家POS+支付+贷款"),
    ("Coupang", "CPNG", "NYSE", "$35B", "S", "电商", "韩国版Amazon"),
    ("CoStar Group", "CSGP", "NASDAQ", "$30B", "S", "商业地产数据", "商业地产数据+市场平台"),
    ("Samsara", "IOT", "NYSE", "$25B", "S", "工业IoT", "工业物联网操作系统"),
    ("Guidewire Software", "GWRE", "NYSE", "$20B", "S", "保险SaaS", "保险行业云核心系统"),
    ("Tyler Technologies", "TYL", "NYSE", "$20B", "S", "政府SaaS", "地方政府全套操作系统"),
    ("MongoDB", "MDB", "NASDAQ", "$20B", "S", "数据库", "开发者数据库平台"),
    ("Toast", "TOST", "NYSE", "$15B", "S", "餐饮SaaS", "餐饮行业操作系统"),
    ("Okta", "OKTA", "NASDAQ", "$15B", "S", "身份认证", "身份认证平台"),
    ("Procore Technologies", "PCOR", "NYSE", "$12B", "S", "建筑SaaS", "建筑施工操作系统"),
    ("Affirm", "AFRM", "NASDAQ", "$12B", "S", "BNPL", "为电商商家提供BNPL"),
    ("Elastic NV", "ESTC", "NYSE", "$10B", "S", "搜索/数据", "搜索+日志+观测性平台"),
    ("Twilio", "TWLO", "NYSE", "$10B", "S", "CPaaS", "API优先通信平台"),
    ("XPO", "XPO", "NYSE", "$10B", "S", "物流", "大件配送+LTL网络"),
    ("ServiceTitan", "TTAN", "NASDAQ", "$9B", "S", "垂直SaaS", "家庭服务业操作系统"),
    ("Global-e Online", "GLBE", "NASDAQ", "$8B", "S", "跨境电商", "跨境电商基础设施"),
    ("monday.com", "MNDY", "NASDAQ", "$8B", "S", "工作管理SaaS", "工作操作系统"),
    ("Shift4 Payments", "FOUR", "NYSE", "$8B", "S", "垂直支付", "餐饮+酒店+娱乐垂直支付"),
    ("AppFolio", "APPF", "NASDAQ", "$8B", "S", "地产SaaS", "地产管理操作系统"),
    ("Global Payments", "GPN", "NYSE", "$8B", "S", "支付", "全球商家收单"),
    ("Descartes Systems", "DSGX", "NASDAQ", "$7.5B", "S", "物流SaaS", "电商物流SaaS"),
    ("Etsy", "ETSY", "NASDAQ", "$7B", "S", "手工电商", "手工/创意卖家平台"),
    ("BILL Holdings", "BILL", "NYSE", "$6B", "S", "金融SaaS", "中小企业财务运营"),
    ("WEX Inc.", "WEX", "NYSE", "$6B", "S", "B2B支付", "B2B支付+企业支出管理"),
    ("Wix.com", "WIX", "NASDAQ", "$5B", "S", "建站/电商SaaS", "建站+电商+支付一站式SaaS"),
    ("GXO Logistics", "GXO", "NYSE", "$5B", "S", "3PL", "纯3PL科技平台"),

    # ── NVIDIA类 (23家) ──
    ("TSMC (台积电)", "TSM", "NYSE", "$1.76T", "N", "晶圆代工", "全球最先进代工"),
    ("Broadcom", "AVGO", "NASDAQ", "$1.6T", "N", "半导体/定制ASIC", "Google TPU共同设计"),
    ("Meta Platforms", "META", "NASDAQ", "$1.5T", "N", "大型科技平台", "MTIA自研推理芯片"),
    ("ASML Holding", "ASML", "NASDAQ", "$250B", "N", "光刻机", "EUV光刻机全球唯一"),
    ("Cisco Systems", "CSCO", "NASDAQ", "$240B", "N", "网络设备", "网络基础设施+AI数据中心"),
    ("Texas Instruments", "TXN", "NASDAQ", "$175B", "N", "模拟IC", "模拟/混合信号IC巨头"),
    ("Advanced Micro Devices", "AMD", "NASDAQ", "$150B", "N", "AI芯片", "GPU+CPU双栈，MI300X"),
    ("Qualcomm", "QCOM", "NASDAQ", "$150B", "N", "半导体", "AI200数据中心+骁龙端侧NPU"),
    ("Arm Holdings", "ARM", "NASDAQ", "$140B", "N", "芯片IP", "AI芯片架构IP"),
    ("Applied Materials", "AMAT", "NASDAQ", "$120B", "N", "半导体设备", "最广谱半导体设备商"),
    ("KLA Corporation", "KLAC", "NASDAQ", "$100B", "N", "过程控制", "过程控制/良率管理全球#1"),
    ("Arista Networks", "ANET", "NYSE", "$100B", "N", "网络设备", "AI数据中心以太网交换机#1"),
    ("Micron Technology", "MU", "NASDAQ", "$100B", "N", "存储", "HBM3E+LPDDR5+NAND"),
    ("ABB Ltd", "ABB", "NYSE", "$100B", "N", "工业自动化", "工业机器人+电气化"),
    ("Intel", "INTC", "NASDAQ", "$90B", "N", "半导体", "Gaudi AI加速器+GPU+FPGA"),
    ("Analog Devices", "ADI", "NASDAQ", "$90B", "N", "模拟IC", "混合信号IC+高速ADC"),
    ("Lam Research", "LRCX", "NASDAQ", "$90B", "N", "刻蚀设备", "刻蚀+清洗设备全球领先"),
    ("Cadence Design Systems", "CDNS", "NASDAQ", "$80B", "N", "EDA", "EDA软件#2+Cerebrus AI"),
    ("Synopsys", "SNPS", "NASDAQ", "$75B", "N", "EDA", "EDA软件#1"),
    ("Marvell Technology", "MRVL", "NASDAQ", "$60B", "N", "半导体", "定制ASIC+光DSP"),
    ("NXP Semiconductors", "NXPI", "NASDAQ", "$50B", "N", "汽车芯片", "汽车MCU/SoC全球#1"),
    ("SMIC (中芯国际)", "0981.HK", "HKEX", "$48B", "N", "晶圆代工", "中国最大晶圆代工"),
    ("CoreWeave", "CRWV", "NASDAQ", "$40B", "N", "GPU云", "AI-Native GPU云"),
]

assert len(DATA) == 100, f"Expected 100, got {len(DATA)}"

# ===== Category helpers =====
def get_fill(cat):
    if len(cat) > 1:
        return MULTI_FILL
    return {"M": META_FILL, "S": SHOP_FILL, "N": NV_FILL}.get(cat, None)

CAT_LABELS = {
    "M": "Meta类", "S": "Shopify类", "N": "NVIDIA类",
    "MS": "Meta+Shopify", "MN": "Meta+NVIDIA", "SN": "Shopify+NVIDIA", "MSN": "三类兼具",
}

# ===== Sheet 1: 主表 =====
ws = wb.active
ws.title = "100家港美股投研标的"

HEADERS = ["序号", "公司名", "Ticker", "交易所", "市值", "分类", "行业", "核心备注"]
COL_WIDTHS = [6, 35, 12, 10, 12, 16, 22, 55]

for i, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for col, h in enumerate(HEADERS, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN_BORDER
ws.row_dimensions[1].height = 22

for i, (name, ticker, exchange, mcap, cat, sector, note) in enumerate(DATA, 1):
    row = i + 1
    cat_label = CAT_LABELS.get(cat, cat)
    vals = [i, name, ticker, exchange, mcap, cat_label, sector, note]
    fill = get_fill(cat)
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = DATA_FONT
        c.border = THIN_BORDER
        c.alignment = Alignment(vertical="center", wrap_text=(col in (7, 8)))
        if fill and col == 6:
            c.fill = fill

ws.auto_filter.ref = f"A1:H{len(DATA)+1}"
ws.freeze_panes = "A2"

# ===== Sheet 2: 汇总 =====
ws2 = wb.create_sheet("汇总统计")
ws2["A1"] = "分类"
ws2["B1"] = "数量（含重叠）"
ws2["C1"] = "说明"
for c in [ws2["A1"], ws2["B1"], ws2["C1"]]:
    c.font = HEADER_FONT
    c.fill = HEADER_FILL

counts = {}
for row in DATA:
    cat = row[4]
    for ch in cat:
        label = CAT_LABELS.get(ch, ch)
        counts[label] = counts.get(label, 0) + 1

descs = {
    "Meta类": "广告/社交/内容/平台型公司",
    "Shopify类": "SaaS/支付/基础设施型公司",
    "NVIDIA类": "半导体/AI芯片/设备/硬件型公司",
}

for i, (k, v) in enumerate(counts.items(), 2):
    ws2.cell(row=i, column=1, value=k).font = DATA_FONT
    ws2.cell(row=i, column=2, value=v).font = DATA_FONT
    ws2.cell(row=i, column=3, value=descs.get(k, "")).font = DATA_FONT

r = len(counts) + 2
ws2.cell(row=r, column=1, value="总计（去重后）").font = Font(name="Arial", bold=True, size=10)
ws2.cell(row=r, column=2, value=len(DATA)).font = Font(name="Arial", bold=True, size=10)
ws2.cell(row=r, column=3, value="部分公司属于多个分类").font = DATA_FONT

ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 35

# ===== 保存 =====
OUT = "/Users/leafsun/Desktop/AI研究/Alike-Investment/100家港美股投研标的.xlsx"
wb.save(OUT)
print(f"✅ Saved to {OUT} ({len(DATA)} companies)")

# Print category breakdown
print(f"\n分类统计（含重叠）:")
for k, v in counts.items():
    print(f"  {k}: {v}家")
