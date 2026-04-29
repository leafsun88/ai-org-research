import json
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch


MODULE_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(MODULE_DIR))


class PublicCompanyCollectionTests(unittest.TestCase):
    def test_public_company_directory_contract(self):
        from public_company.common import PublicCompanyTarget, ensure_public_company_dirs

        with tempfile.TemporaryDirectory() as tmp:
            target = PublicCompanyTarget(
                ticker="SNDK",
                market="US",
                slug="sandisk",
                company_name="Sandisk",
                company_root=Path(tmp) / "companies" / "sandisk",
            )
            dirs = ensure_public_company_dirs(target)
            root = target.company_root / "vault" / "public_company"

            expected = [
                "raw_filings/annual_reports",
                "raw_filings/quarterly_reports",
                "raw_filings/prospectus",
                "raw_filings/earnings_calls",
                "raw_filings/governance",
                "raw_filings/material_events",
                "processed/mda",
                "processed/earnings_calls",
                "processed/books",
                "processed/source_notes",
                "data/statements",
                "data/market_data",
                "data/valuation",
                "data/capital_signals",
                "semi_public/investor_views",
                "semi_public/short_seller_reports",
                "metadata/probes",
            ]
            self.assertTrue(all((root / item).is_dir() for item in expected))
            self.assertIn(root / "metadata" / "collection_status.json", dirs["metadata_files"])

    def test_sec_filing_classification(self):
        from public_company.sec_us import classify_filing

        self.assertEqual(classify_filing("10-K", ""), "annual_reports")
        self.assertEqual(classify_filing("ARS", "ANNUAL REPORT TO SECURITY HOLDERS"), "annual_reports")
        self.assertEqual(classify_filing("10-Q", ""), "quarterly_reports")
        self.assertEqual(classify_filing("S-1", ""), "prospectus")
        self.assertEqual(classify_filing("F-1", ""), "prospectus")
        self.assertEqual(classify_filing("424B4", "Prospectus"), "prospectus")
        self.assertEqual(classify_filing("DEF 14A", ""), "governance")
        self.assertEqual(classify_filing("8-K", "Results of Operations"), "earnings_calls")
        self.assertEqual(classify_filing("8-K", "Entry into a Material Agreement"), "material_events")

    def test_mda_extraction_from_10k_text(self):
        from public_company.sec_us import extract_mda_section

        text = """
        ITEM 6. RESERVED
        ITEM 7. MANAGEMENT'S DISCUSSION AND ANALYSIS OF FINANCIAL CONDITION AND RESULTS OF OPERATIONS
        Revenue increased as demand changed. Gross margin improved.
        We discuss liquidity and capital resources here.
        ITEM 7A. QUANTITATIVE AND QUALITATIVE DISCLOSURES ABOUT MARKET RISK
        Market risk text.
        ITEM 8. FINANCIAL STATEMENTS AND SUPPLEMENTARY DATA
        """
        section = extract_mda_section(text)

        self.assertIn("Revenue increased", section)
        self.assertIn("liquidity", section)
        self.assertNotIn("Market risk text", section)

    def test_mda_extraction_skips_table_of_contents_and_item_references(self):
        from public_company.sec_us import extract_mda_section

        text = """
        Item 7. Management's Discussion and Analysis of Financial Condition and Results of Operations 45
        Item 7A. Quantitative and Qualitative Disclosures About Market Risk 56
        44 Table of Contents
        Item 7.     Management's Discussion and Analysis of Financial Condition and Results of Operations
        The following discussion and analysis should be read with Part II, Item 8., Financial Statements.
        Results of Operations Overview
        Revenue grew because pricing improved and demand recovered.
        Liquidity and capital resources remained adequate.
        55 Table of Contents
        Item 7A.     Quantitative and Qualitative Disclosures About Market Risk
        Foreign currency risk text.
        """
        section = extract_mda_section(text)

        self.assertIn("Results of Operations Overview", section)
        self.assertIn("Revenue grew", section)
        self.assertNotIn("Foreign currency risk text", section)

    def test_financial_model_workbook_contract(self):
        from openpyxl import load_workbook
        from public_company.common import PublicCompanyTarget, ensure_public_company_dirs
        from public_company.financial_model import write_financial_model

        with tempfile.TemporaryDirectory() as tmp:
            target = PublicCompanyTarget(
                ticker="SNDK",
                market="US",
                slug="sandisk",
                company_name="Sandisk",
                company_root=Path(tmp) / "companies" / "sandisk",
                peer_tickers=["WDC", "MU"],
            )
            ensure_public_company_dirs(target)
            root = target.company_root / "vault" / "public_company"
            statements_dir = root / "data" / "statements"
            (statements_dir / "annual_statements.json").write_text(
                json.dumps(
                    {
                        "periods": [
                            {
                                "period": "2025",
                                "revenue": 8700000000,
                                "gross_profit": 2100000000,
                                "operating_income": 450000000,
                                "net_income": 320000000,
                                "capital_expenditures": -800000000,
                                "operating_cash_flow": 1300000000,
                                "cash": 1600000000,
                                "debt": 4200000000,
                                "shares_outstanding": 300000000,
                            }
                        ],
                        "source_map": {"revenue": "sample_companyfacts.json"},
                    }
                ),
                encoding="utf-8",
            )
            (root / "data" / "market_data" / "price_history.csv").write_text(
                "date,open,high,low,close,adj_close,volume\n2026-04-24,50,52,49,51,51,1000000\n",
                encoding="utf-8",
            )
            (root / "data" / "valuation" / "valuation_snapshot.json").write_text(
                json.dumps({"market_cap": 15300000000, "enterprise_value": 17900000000}),
                encoding="utf-8",
            )

            result = write_financial_model(target, as_of_date="2026-04-25")
            workbook = Path(result["workbook_path"])
            validation = json.loads(Path(result["validation_path"]).read_text(encoding="utf-8"))
            wb = load_workbook(workbook, data_only=False)

            self.assertEqual(
                wb.sheetnames,
                [
                    "Summary",
                    "Historical Statements",
                    "Operating Metrics",
                    "Forecast",
                    "DCF",
                    "Comps",
                    "Sensitivity",
                    "Capital Signals",
                    "Source Map",
                    "Assumptions",
                ],
            )
            self.assertTrue(validation["checks"]["workbook_created"]["passed"])
            self.assertTrue(validation["checks"]["required_sheets"]["passed"])
            self.assertFalse(validation["fatal_errors"])
            self.assertTrue(validation["checks"]["operating_metrics_visible_values"]["passed"])
            self.assertEqual(validation["status"], "needs_review")
            self.assertIn("quarterly_income_statement", validation["gaps"])
            self.assertIn("comps_multiples", validation["gaps"])

    def test_model_validation_does_not_pass_on_low_confidence_market_cap(self):
        from public_company.common import PublicCompanyTarget, ensure_public_company_dirs
        from public_company.financial_model import write_financial_model

        with tempfile.TemporaryDirectory() as tmp:
            target = PublicCompanyTarget(
                ticker="SNDK",
                market="US",
                slug="sandisk",
                company_name="Sandisk",
                company_root=Path(tmp) / "companies" / "sandisk",
                peer_tickers=["WDC"],
            )
            ensure_public_company_dirs(target)
            root = target.company_root / "vault" / "public_company"
            (root / "data" / "statements" / "annual_statements.json").write_text(
                json.dumps(
                    {
                        "periods": [
                            {
                                "period": "2025",
                                "revenue": 1000,
                                "gross_profit": 500,
                                "operating_income": 100,
                                "net_income": 80,
                                "operating_cash_flow": 90,
                                "capital_expenditures": -20,
                                "cash": 50,
                                "debt": 10,
                                "assets": 200,
                                "liabilities": 90,
                                "stockholders_equity": 110,
                                "shares_outstanding": 100,
                            }
                        ],
                        "source_map": {},
                    }
                ),
                encoding="utf-8",
            )
            (root / "data" / "statements" / "quarterly_statements.json").write_text(
                json.dumps({"periods": [{"period": "2026-Q1", "revenue": 250, "net_income": 20}]}),
                encoding="utf-8",
            )
            (root / "data" / "market_data" / "price_history.csv").write_text(
                "date,open,high,low,close,adj_close,volume\n2026-04-24,10,10,10,10,10,100\n",
                encoding="utf-8",
            )
            (root / "data" / "valuation" / "valuation_snapshot.json").write_text(
                json.dumps(
                    {
                        "market_cap": 1000,
                        "enterprise_value": 960,
                        "market_cap_source": "computed_from_weighted_average_diluted_shares",
                        "market_cap_confidence": "low",
                    }
                ),
                encoding="utf-8",
            )
            (root / "data" / "valuation" / "peer_valuation.json").write_text(
                json.dumps({"peers": [{"ticker": "WDC", "price_to_sales": 2.0, "ev_to_ebitda": 8.0}]}),
                encoding="utf-8",
            )

            result = write_financial_model(target, as_of_date="2026-04-25")
            validation = json.loads(Path(result["validation_path"]).read_text(encoding="utf-8"))

            self.assertEqual(validation["status"], "needs_review")
            self.assertIn("market_cap_confidence", validation["gaps"])

    def test_companyfacts_parser_prefers_latest_period_and_true_quarter(self):
        from public_company.common import PublicCompanyTarget, ensure_public_company_dirs
        from public_company.sec_us import parse_companyfacts_to_statements

        def fact(val, fy, fp, form, filed, end, start=None, frame=None):
            payload = {"val": val, "fy": fy, "fp": fp, "form": form, "filed": filed, "end": end}
            if start:
                payload["start"] = start
            if frame:
                payload["frame"] = frame
            return payload

        def usd(entries):
            return {"units": {"USD": entries}}

        with tempfile.TemporaryDirectory() as tmp:
            target = PublicCompanyTarget(
                ticker="SNDK",
                market="US",
                slug="sandisk",
                company_name="Sandisk",
                company_root=Path(tmp) / "companies" / "sandisk",
            )
            ensure_public_company_dirs(target)
            facts = {
                "facts": {
                    "us-gaap": {
                        "RevenueFromContractWithCustomerExcludingAssessedTax": usd(
                            [
                                fact(6086, 2025, "FY", "10-K", "2025-08-29", "2023-06-30", "2022-07-01"),
                                fact(7355, 2025, "FY", "10-K", "2025-08-29", "2025-06-27", "2024-06-29"),
                                fact(5333, 2026, "Q2", "10-Q", "2026-01-30", "2026-01-02", "2025-06-28"),
                                fact(3025, 2026, "Q2", "10-Q", "2026-01-30", "2026-01-02", "2025-10-04", "CY2025Q4"),
                            ]
                        ),
                        "GrossProfit": usd([fact(1736, 2025, "FY", "10-K", "2025-08-29", "2025-06-27", "2024-06-29")]),
                        "OperatingIncomeLoss": usd(
                            [
                                fact(-1377, 2025, "FY", "10-K", "2025-08-29", "2025-06-27", "2024-06-29"),
                                fact(1065, 2026, "Q2", "10-Q", "2026-01-30", "2026-01-02", "2025-10-04", "CY2025Q4"),
                            ]
                        ),
                        "NetIncomeLoss": usd(
                            [
                                fact(-1641, 2025, "FY", "10-K", "2025-08-29", "2025-06-27", "2024-06-29"),
                                fact(803, 2026, "Q2", "10-Q", "2026-01-30", "2026-01-02", "2025-10-04", "CY2025Q4"),
                            ]
                        ),
                        "NetCashProvidedByUsedInOperatingActivities": usd([fact(-846, 2025, "FY", "10-K", "2025-08-29", "2025-06-27", "2024-06-29")]),
                        "PaymentsToAcquirePropertyPlantAndEquipment": usd([fact(139, 2025, "FY", "10-K", "2025-08-29", "2025-06-27", "2024-06-29")]),
                        "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents": usd([fact(1481, 2025, "FY", "10-K", "2025-08-29", "2025-06-27")]),
                        "Assets": usd([fact(12985, 2025, "FY", "10-K", "2025-08-29", "2025-06-27")]),
                        "Liabilities": usd([fact(3769, 2025, "FY", "10-K", "2025-08-29", "2025-06-27")]),
                        "StockholdersEquity": usd([fact(9216, 2025, "FY", "10-K", "2025-08-29", "2025-06-27")]),
                        "LongTermDebtCurrent": usd([fact(20, 2025, "FY", "10-K", "2025-08-29", "2025-06-27")]),
                        "LongTermDebtNoncurrent": usd([fact(1829, 2025, "FY", "10-K", "2025-08-29", "2025-06-27")]),
                        "WeightedAverageNumberOfDilutedSharesOutstanding": {
                            "units": {"shares": [fact(145, 2025, "FY", "10-K", "2025-08-29", "2025-06-27", "2024-06-29")]}
                        },
                    }
                }
            }
            (target.public_company_root / "metadata" / "sec_companyfacts.json").write_text(json.dumps(facts), encoding="utf-8")

            parse_companyfacts_to_statements(target)
            annual = json.loads((target.public_company_root / "data" / "statements" / "annual_statements.json").read_text())
            quarterly = json.loads((target.public_company_root / "data" / "statements" / "quarterly_statements.json").read_text())

            latest = annual["periods"][-1]
            q2 = [row for row in quarterly["periods"] if row["period"] == "2026-Q2"][0]
            self.assertEqual(latest["revenue"], 7355)
            self.assertEqual(latest["assets"], 12985)
            self.assertEqual(latest["debt"], 1849.0)
            self.assertEqual(latest["capital_expenditures"], -139)
            self.assertEqual(q2["revenue"], 3025)
            self.assertEqual(q2["net_income"], 803)

    def test_quality_loop_repairs_then_passes(self):
        from public_company.common import PublicCompanyTarget
        from public_company.quality_loop import run_quality_loop

        target = PublicCompanyTarget(
            ticker="SNDK",
            market="US",
            slug="sandisk",
            company_name="Sandisk",
            company_root=Path(tempfile.mkdtemp()) / "companies" / "sandisk",
        )
        with patch("public_company.quality_loop.write_financial_model", return_value={"workbook_path": "model.xlsx"}), patch(
            "public_company.quality_loop.audit_public_company",
            side_effect=[
                {"status": "needs_review", "gaps": ["annual_statement_completeness"]},
                {"status": "pass", "gaps": []},
            ],
        ), patch("public_company.quality_loop.parse_companyfacts_to_statements", return_value={"status": "ok"}) as repair, patch(
            "public_company.quality_loop.discover_source_strategy", return_value={"status": "ok"}
        ):
            result = run_quality_loop(target, max_iterations=2, as_of_date="2026-04-25")

        self.assertEqual(result["status"], "pass")
        self.assertTrue(repair.called)

    def test_quality_loop_blocks_when_gaps_do_not_change(self):
        from public_company.common import PublicCompanyTarget
        from public_company.quality_loop import run_quality_loop

        target = PublicCompanyTarget(
            ticker="SNDK",
            market="US",
            slug="sandisk",
            company_name="Sandisk",
            company_root=Path(tempfile.mkdtemp()) / "companies" / "sandisk",
        )
        with patch("public_company.quality_loop.write_financial_model", return_value={"workbook_path": "model.xlsx"}), patch(
            "public_company.quality_loop.audit_public_company",
            return_value={"status": "needs_review", "gaps": ["earnings_transcripts"]},
        ), patch("public_company.quality_loop.discover_source_strategy", return_value={"status": "ok"}):
            result = run_quality_loop(target, max_iterations=2, as_of_date="2026-04-25")

            self.assertEqual(result["status"], "blocked")
            self.assertIn("earnings_transcripts", result["remaining_gaps"])
            self.assertIn("earnings_transcripts", result["blocker_notes"])

    def test_audit_flags_low_density_statement_rows(self):
        from public_company.audit import audit_public_company
        from public_company.common import PublicCompanyTarget, ensure_public_company_dirs

        with tempfile.TemporaryDirectory() as tmp:
            target = PublicCompanyTarget(
                ticker="SNDK",
                market="US",
                slug="sandisk",
                company_name="Sandisk",
                company_root=Path(tmp) / "companies" / "sandisk",
            )
            ensure_public_company_dirs(target)
            root = target.public_company_root
            (root / "data" / "statements" / "annual_statements.json").write_text(
                json.dumps({"periods": [{"period": "2025", "revenue": 1000, "gross_profit": None, "operating_income": None}]}),
                encoding="utf-8",
            )
            (root / "data" / "statements" / "quarterly_statements.json").write_text(
                json.dumps({"periods": [{"period": "2026-Q1", "revenue": 250, "gross_profit": 100, "operating_income": 20, "net_income": 10, "assets": 500, "liabilities": 200, "cash": 50}]}),
                encoding="utf-8",
            )

            audit = audit_public_company(target, model_validation={"status": "pass", "gaps": []})

            self.assertIn("annual_statement_density", audit["gaps"])
            self.assertFalse(audit["checks"]["annual_statement_density"]["passed"])

    def test_audit_flags_wrong_statement_json_schema(self):
        from public_company.audit import audit_public_company
        from public_company.common import PublicCompanyTarget, ensure_public_company_dirs

        with tempfile.TemporaryDirectory() as tmp:
            target = PublicCompanyTarget(
                ticker="SNDK",
                market="US",
                slug="sandisk",
                company_name="Sandisk",
                company_root=Path(tmp) / "companies" / "sandisk",
            )
            ensure_public_company_dirs(target)
            root = target.public_company_root
            (root / "data" / "statements" / "annual_statements.json").write_text(
                json.dumps({"annual_statements": [{"period": "2025", "revenue": 1000}]}),
                encoding="utf-8",
            )
            (root / "data" / "statements" / "quarterly_statements.json").write_text(
                json.dumps({"quarterly_statements": [{"period": "2026-Q1", "revenue": 250}]}),
                encoding="utf-8",
            )

            audit = audit_public_company(target, model_validation={"status": "pass", "gaps": []})

            self.assertIn("annual_statement_schema", audit["gaps"])
            self.assertIn("quarterly_statement_schema", audit["gaps"])
            self.assertFalse(audit["checks"]["annual_statement_schema"]["passed"])
            self.assertFalse(audit["checks"]["quarterly_statement_schema"]["passed"])

    def test_audit_requires_annual_report_pdf_when_collecting_raw_filings(self):
        from public_company.audit import audit_public_company
        from public_company.common import PublicCompanyTarget, ensure_public_company_dirs

        with tempfile.TemporaryDirectory() as tmp:
            target = PublicCompanyTarget(
                ticker="SNDK",
                market="US",
                slug="sandisk",
                company_name="Sandisk",
                company_root=Path(tmp) / "companies" / "sandisk",
            )
            ensure_public_company_dirs(target)
            root = target.public_company_root
            (root / "raw_filings" / "annual_reports" / "sample_10k.htm").write_text(
                "Sandisk Corporation Form 10-K Annual Report", encoding="utf-8"
            )
            (root / "data" / "statements" / "annual_statements.json").write_text(json.dumps({"periods": []}), encoding="utf-8")
            (root / "data" / "statements" / "quarterly_statements.json").write_text(json.dumps({"periods": []}), encoding="utf-8")

            audit = audit_public_company(target, model_validation={"status": "pass", "gaps": []})

            self.assertIn("annual_report_pdf", audit["gaps"])
            self.assertFalse(audit["checks"]["annual_report_pdf"]["passed"])


if __name__ == "__main__":
    unittest.main()
