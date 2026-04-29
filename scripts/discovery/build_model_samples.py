"""Build sample public-company model workbooks for comparing modeling styles."""

from __future__ import annotations

import argparse
import json
import statistics
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

MODULE_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = MODULE_DIR.parents[1]
sys.path.insert(0, str(MODULE_DIR))

from target_config import load_target  # noqa: E402
from public_company.common import PublicCompanyTarget, TODAY, read_json, write_json  # noqa: E402


NAVY = "1F4E79"
BLUE = "5B9BD5"
LT_BLUE = "D9EAF7"
GREY = "F2F2F2"
LT_GREY = "F7F7F7"
GREEN = "E2F0D9"
YELLOW = "FFF2CC"
WHITE = "FFFFFF"
RED = "F4CCCC"

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _num(value: Any, default: float = 0.0) -> float:
    if value in {None, ""}:
        return default
    if isinstance(value, str):
        value = value.replace("$", "").replace(",", "").replace("%", "")
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _ratio(numerator: Any, denominator: Any) -> float | None:
    denominator_f = _num(denominator)
    if denominator_f == 0:
        return None
    return _num(numerator) / denominator_f


def _median(values: list[float]) -> float | None:
    clean = [value for value in values if value is not None]
    return statistics.median(clean) if clean else None


def _load_inputs(target: PublicCompanyTarget) -> dict[str, Any]:
    root = target.public_company_root
    return {
        "annual": read_json(root / "data" / "statements" / "annual_statements.json", {"periods": []}) or {"periods": []},
        "quarterly": read_json(root / "data" / "statements" / "quarterly_statements.json", {"periods": []}) or {"periods": []},
        "valuation": read_json(root / "data" / "valuation" / "valuation_snapshot.json", {}) or {},
        "peers": read_json(root / "data" / "valuation" / "peer_valuation.json", {"peers": []}) or {"peers": []},
        "holders": read_json(root / "data" / "capital_signals" / "institutional_holders.json", {"holders": []}) or {"holders": []},
    }


def _latest(inputs: dict[str, Any]) -> dict[str, Any]:
    periods = inputs["annual"].get("periods", [])
    return periods[-1] if periods else {}


def _apply_base_style(wb: Workbook) -> None:
    for ws in wb.worksheets:
        ws.sheet_view.zoomScale = 90
        ws.freeze_panes = "A2"
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = BORDER
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0;(#,##0);-'
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True, color=WHITE)
                cell.fill = PatternFill("solid", fgColor=NAVY)
        for column_cells in ws.columns:
            values = [str(cell.value) for cell in column_cells if cell.value is not None]
            width = min(max(max((len(value) for value in values), default=10) + 2, 12), 32)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def _set_percent_columns(ws, cols: list[int], start_row: int = 2) -> None:
    for col in cols:
        for row in range(start_row, ws.max_row + 1):
            ws.cell(row, col).number_format = "0.0%;(0.0%);-"


def _set_multiple_columns(ws, cols: list[int], start_row: int = 2) -> None:
    for col in cols:
        for row in range(start_row, ws.max_row + 1):
            ws.cell(row, col).number_format = "0.0x;(0.0x);-"


def _fill_row(ws, row: int, color: str, font_color: str = "000000", bold: bool = False) -> None:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(color=font_color, bold=bold)


SCENARIOS = {
    "Bear": {
        "revenue_growth": 0.00,
        "operating_margin": 0.08,
        "tax_rate": 0.21,
        "da_percent_revenue": 0.04,
        "capex_percent_revenue": 0.09,
        "nwc_percent_revenue": 0.025,
        "wacc": 0.12,
        "terminal_growth": 0.02,
    },
    "Base": {
        "revenue_growth": 0.05,
        "operating_margin": 0.12,
        "tax_rate": 0.21,
        "da_percent_revenue": 0.04,
        "capex_percent_revenue": 0.08,
        "nwc_percent_revenue": 0.02,
        "wacc": 0.10,
        "terminal_growth": 0.03,
    },
    "Bull": {
        "revenue_growth": 0.10,
        "operating_margin": 0.18,
        "tax_rate": 0.21,
        "da_percent_revenue": 0.04,
        "capex_percent_revenue": 0.07,
        "nwc_percent_revenue": 0.015,
        "wacc": 0.09,
        "terminal_growth": 0.035,
    },
}


def _dcf(latest: dict[str, Any], assumptions: dict[str, float]) -> dict[str, Any]:
    revenue = _num(latest.get("revenue"))
    cash = _num(latest.get("cash"))
    debt = _num(latest.get("debt"))
    shares = _num(latest.get("shares_outstanding"))
    growth = assumptions["revenue_growth"]
    operating_margin = assumptions["operating_margin"]
    tax_rate = assumptions["tax_rate"]
    da_percent = assumptions["da_percent_revenue"]
    capex_percent = assumptions["capex_percent_revenue"]
    nwc_percent = assumptions["nwc_percent_revenue"]
    wacc = assumptions["wacc"]
    terminal_growth = assumptions["terminal_growth"]
    forecast: list[dict[str, float]] = []
    current_revenue = revenue
    for year in range(1, 6):
        current_revenue *= 1 + growth
        operating_income = current_revenue * operating_margin
        nopat = operating_income * (1 - tax_rate)
        da = current_revenue * da_percent
        capex = -current_revenue * capex_percent
        nwc = -current_revenue * nwc_percent
        fcf = nopat + da + capex + nwc
        forecast.append(
            {
                "year": year,
                "revenue": current_revenue,
                "operating_income": operating_income,
                "nopat": nopat,
                "da": da,
                "capex": capex,
                "nwc": nwc,
                "fcf": fcf,
                "pv_fcf": fcf / ((1 + wacc) ** year),
            }
        )
    if wacc <= terminal_growth:
        return {"status": "invalid", "forecast": forecast}
    terminal_fcf = forecast[-1]["fcf"] * (1 + terminal_growth)
    terminal_value = terminal_fcf / (wacc - terminal_growth)
    pv_terminal = terminal_value / ((1 + wacc) ** 5)
    pv_fcf = sum(row["pv_fcf"] for row in forecast)
    enterprise_value = pv_fcf + pv_terminal
    equity_value = enterprise_value + cash - debt
    implied_price = equity_value / shares if shares else None
    terminal_share = pv_terminal / enterprise_value if enterprise_value else None
    return {
        "status": "calculated",
        "forecast": forecast,
        "pv_fcf": pv_fcf,
        "terminal_value": terminal_value,
        "pv_terminal_value": pv_terminal,
        "enterprise_value": enterprise_value,
        "equity_value": equity_value,
        "implied_price": implied_price,
        "terminal_value_share": terminal_share,
    }


def _write_sources(ws, target: PublicCompanyTarget) -> None:
    ws.append(["Field", "Source", "Confidence"])
    rows = [
        ("Financial statements", "data/statements/annual_statements.json; SEC companyfacts-derived", "medium-high"),
        ("Quarterly statements", "data/statements/quarterly_statements.json; SEC companyfacts-derived", "medium"),
        ("Market data", "data/valuation/valuation_snapshot.json; public fallback", "medium"),
        ("Peer multiples", "data/valuation/peer_valuation.json; public fallback", "medium"),
        ("Capital signals", "data/capital_signals/institutional_holders.json", "medium"),
        ("Company config", f"config/company_targets.json::{target.ticker}", "high"),
    ]
    for row in rows:
        ws.append(row)


def build_compact_model(target: PublicCompanyTarget, inputs: dict[str, Any], out_dir: Path, as_of: str) -> Path:
    latest = _latest(inputs)
    valuation = inputs["valuation"]
    base = _dcf(latest, SCENARIOS["Base"])
    peers = inputs["peers"].get("peers", [])
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Item", "Value", "Source"])
    rows = [
        ("Company", target.company_name, "target config"),
        ("Ticker", target.ticker, "target config"),
        ("As of", as_of, "script"),
        ("Current price", valuation.get("last_price"), "valuation_snapshot.json"),
        ("Market cap", valuation.get("market_cap"), "valuation_snapshot.json"),
        ("Enterprise value", valuation.get("enterprise_value"), "valuation_snapshot.json"),
        ("Latest revenue", latest.get("revenue"), "annual_statements.json"),
        ("Latest net income", latest.get("net_income"), "annual_statements.json"),
        ("Base DCF EV", base.get("enterprise_value"), "DCF"),
        ("Base DCF equity value", base.get("equity_value"), "DCF"),
        ("Base DCF implied price", base.get("implied_price"), "DCF"),
        ("Model style", "Compact DCF + comps", "SkillsMP review"),
    ]
    for row in rows:
        ws.append(row)
    ws = wb.create_sheet("Historical")
    ws.append(["period", "revenue", "gross_margin", "operating_margin", "net_margin", "fcf_margin", "assets", "cash", "debt"])
    for row in inputs["annual"].get("periods", []):
        ws.append(
            [
                row.get("period"),
                row.get("revenue"),
                _ratio(row.get("gross_profit"), row.get("revenue")),
                _ratio(row.get("operating_income"), row.get("revenue")),
                _ratio(row.get("net_income"), row.get("revenue")),
                _ratio(_num(row.get("operating_cash_flow")) + _num(row.get("capital_expenditures")), row.get("revenue")),
                row.get("assets"),
                row.get("cash"),
                row.get("debt"),
            ]
        )
    _set_percent_columns(ws, [3, 4, 5, 6])
    ws = wb.create_sheet("DCF")
    ws.append(["Metric", "Y1", "Y2", "Y3", "Y4", "Y5"])
    for metric in ["revenue", "operating_income", "nopat", "da", "capex", "nwc", "fcf", "pv_fcf"]:
        ws.append([metric, *[row.get(metric) for row in base.get("forecast", [])]])
    ws.append(["PV terminal value", base.get("pv_terminal_value")])
    ws.append(["Enterprise value", base.get("enterprise_value")])
    ws.append(["Equity value", base.get("equity_value")])
    ws.append(["Implied price", base.get("implied_price")])
    ws.append(["Terminal value share", base.get("terminal_value_share")])
    ws = wb.create_sheet("Comps")
    ws.append(["ticker", "price", "market_cap", "enterprise_value", "P/S", "EV/EBITDA", "P/E", "source"])
    for peer in peers:
        ws.append(
            [
                peer.get("ticker"),
                peer.get("last_price"),
                peer.get("market_cap"),
                peer.get("enterprise_value"),
                peer.get("price_to_sales"),
                peer.get("ev_to_ebitda"),
                peer.get("trailing_pe"),
                peer.get("source_url") or peer.get("source"),
            ]
        )
    _set_multiple_columns(ws, [5, 6, 7])
    ws = wb.create_sheet("Sources")
    _write_sources(ws, target)
    _apply_base_style(wb)
    path = out_dir / f"{target.ticker.lower()}_compact_dcf_comps.xlsx"
    wb.save(path)
    return path


def build_scenario_model(target: PublicCompanyTarget, inputs: dict[str, Any], out_dir: Path, as_of: str) -> Path:
    latest = _latest(inputs)
    valuation = inputs["valuation"]
    results = {name: _dcf(latest, assumptions) for name, assumptions in SCENARIOS.items()}
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.append(["Metric", "Bear", "Base", "Bull", "Source"])
    for label, key in [
        ("DCF enterprise value", "enterprise_value"),
        ("DCF equity value", "equity_value"),
        ("DCF implied price", "implied_price"),
        ("Terminal value share", "terminal_value_share"),
    ]:
        ws.append([label, results["Bear"].get(key), results["Base"].get(key), results["Bull"].get(key), "DCF by scenario"])
    ws.append(["Current price", valuation.get("last_price"), valuation.get("last_price"), valuation.get("last_price"), "valuation_snapshot.json"])
    ws.append(["Current market cap", valuation.get("market_cap"), valuation.get("market_cap"), valuation.get("market_cap"), "valuation_snapshot.json"])
    ws.append(["Style", "Scenario Architect", "Scenario Architect", "Scenario Architect", "SkillsMP review"])
    _fill_row(ws, 1, NAVY, WHITE, True)
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value in {"DCF implied price", "Style"}:
            _fill_row(ws, row, GREEN if ws.cell(row, 1).value != "Style" else LT_BLUE, "000000", True)
    ws = wb.create_sheet("Assumptions")
    ws.append(["Assumption", "Bear", "Base", "Bull", "Source", "Status"])
    for key in SCENARIOS["Base"]:
        ws.append([key, SCENARIOS["Bear"][key], SCENARIOS["Base"][key], SCENARIOS["Bull"][key], "auto_default; editable", "draft"])
    _set_percent_columns(ws, [2, 3, 4])
    ws = wb.create_sheet("Historical")
    ws.append(["period", "revenue", "gross_profit", "operating_income", "net_income", "operating_cash_flow", "capex"])
    for row in inputs["annual"].get("periods", []):
        ws.append([row.get("period"), row.get("revenue"), row.get("gross_profit"), row.get("operating_income"), row.get("net_income"), row.get("operating_cash_flow"), row.get("capital_expenditures")])
    ws = wb.create_sheet("Forecast")
    ws.append(["Scenario", "Metric", "Y1", "Y2", "Y3", "Y4", "Y5"])
    for name, result in results.items():
        for metric in ["revenue", "operating_income", "fcf"]:
            ws.append([name, metric, *[row.get(metric) for row in result.get("forecast", [])]])
    ws = wb.create_sheet("DCF")
    ws.append(["Scenario", "PV FCF", "PV terminal", "EV", "Equity value", "Implied price", "Terminal value share"])
    for name, result in results.items():
        ws.append([name, result.get("pv_fcf"), result.get("pv_terminal_value"), result.get("enterprise_value"), result.get("equity_value"), result.get("implied_price"), result.get("terminal_value_share")])
    _set_percent_columns(ws, [7])
    ws = wb.create_sheet("Sensitivity")
    ws.append(["WACC \\ terminal growth", "2.0%", "2.5%", "3.0%", "3.5%", "4.0%"])
    for wacc in [0.08, 0.09, 0.10, 0.11, 0.12]:
        row = [wacc]
        for tg in [0.02, 0.025, 0.03, 0.035, 0.04]:
            assumptions = dict(SCENARIOS["Base"], wacc=wacc, terminal_growth=tg)
            row.append(_dcf(latest, assumptions).get("implied_price"))
        ws.append(row)
    _set_percent_columns(ws, [1])
    ws = wb.create_sheet("Checks")
    ws.append(["Check", "Status", "Details"])
    base = results["Base"]
    tv_share = base.get("terminal_value_share")
    checks = [
        ("terminal_value_share_reasonable", "warning" if tv_share is not None and (tv_share > 0.9 or tv_share < 0.4) else "pass", tv_share),
        ("scenario_count", "pass", len(SCENARIOS)),
        ("source_backed_latest_revenue", "pass" if latest.get("revenue") else "fail", latest.get("revenue")),
        ("market_data_available", "pass" if valuation.get("market_cap") else "fail", valuation.get("market_cap")),
    ]
    for row in checks:
        ws.append(row)
    ws = wb.create_sheet("Sources")
    _write_sources(ws, target)
    _apply_base_style(wb)
    path = out_dir / f"{target.ticker.lower()}_scenario_architect.xlsx"
    wb.save(path)
    return path


def build_comps_model(target: PublicCompanyTarget, inputs: dict[str, Any], out_dir: Path, as_of: str) -> Path:
    latest = _latest(inputs)
    valuation = inputs["valuation"]
    peers = [peer for peer in inputs["peers"].get("peers", []) if peer.get("price_to_sales") or peer.get("ev_to_ebitda") or peer.get("trailing_pe")]
    ps_values = [_num(peer.get("price_to_sales"), None) for peer in peers if peer.get("price_to_sales") is not None]
    ev_ebitda_values = [_num(peer.get("ev_to_ebitda"), None) for peer in peers if peer.get("ev_to_ebitda") is not None]
    median_ps = _median(ps_values)
    median_ev_ebitda = _median(ev_ebitda_values)
    revenue = _num(latest.get("revenue"))
    base_ev = revenue * median_ps if median_ps is not None else None
    wb = Workbook()
    ws = wb.active
    ws.title = "Comps Dashboard"
    ws.append(["Metric", "Value", "Source"])
    ws.append(["Target ticker", target.ticker, "target config"])
    ws.append(["Target current EV", valuation.get("enterprise_value"), "valuation_snapshot.json"])
    ws.append(["Target current P/S", valuation.get("price_to_sales"), "valuation_snapshot.json"])
    ws.append(["Peer median P/S", median_ps, "Peer Table"])
    ws.append(["Peer median EV/EBITDA", median_ev_ebitda, "Peer Table"])
    ws.append(["Revenue-implied EV at median P/S", base_ev, "Latest revenue x peer median P/S"])
    ws.append(["Style", "Comps-focused model", "SkillsMP review"])
    ws = wb.create_sheet("Peer Table")
    ws.append(["ticker", "price", "market_cap", "enterprise_value", "P/S", "EV/EBITDA", "P/E", "source"])
    ws.append([target.ticker, valuation.get("last_price"), valuation.get("market_cap"), valuation.get("enterprise_value"), valuation.get("price_to_sales"), valuation.get("ev_to_ebitda"), valuation.get("trailing_pe"), "valuation_snapshot.json"])
    for peer in peers:
        ws.append([peer.get("ticker"), peer.get("last_price"), peer.get("market_cap"), peer.get("enterprise_value"), peer.get("price_to_sales"), peer.get("ev_to_ebitda"), peer.get("trailing_pe"), peer.get("source_url") or peer.get("source")])
    stats_start = ws.max_row + 2
    for label, values in [
        ("Maximum", ps_values),
        ("75th Percentile", sorted(ps_values)[int((len(ps_values) - 1) * 0.75)] if ps_values else []),
        ("Median", median_ps),
        ("Minimum", ps_values),
    ]:
        if isinstance(values, list):
            value = max(values) if label == "Maximum" and values else min(values) if label == "Minimum" and values else None
        else:
            value = values
        ws.append([label, None, None, None, value, None, None, None])
    for row in range(stats_start, ws.max_row + 1):
        _fill_row(ws, row, GREY, "000000", True)
    _set_multiple_columns(ws, [5, 6, 7])
    ws = wb.create_sheet("Implied Value")
    ws.append(["Method", "Low", "Base", "High", "Notes"])
    if median_ps is not None:
        ws.append(["P/S", revenue * median_ps * 0.8, revenue * median_ps, revenue * median_ps * 1.2, "Uses peer median P/S +/- 20%"])
    if median_ev_ebitda is not None:
        ebitda_proxy = _num(latest.get("operating_income")) + revenue * 0.04
        ws.append(["EV/EBITDA", ebitda_proxy * median_ev_ebitda * 0.8, ebitda_proxy * median_ev_ebitda, ebitda_proxy * median_ev_ebitda * 1.2, "EBITDA proxy = EBIT + D&A as 4% revenue"])
    ws = wb.create_sheet("Sources")
    _write_sources(ws, target)
    _apply_base_style(wb)
    path = out_dir / f"{target.ticker.lower()}_comps_focus.xlsx"
    wb.save(path)
    return path


def build_challenger_pack(target: PublicCompanyTarget, inputs: dict[str, Any], out_dir: Path, as_of: str) -> Path:
    latest = _latest(inputs)
    valuation = inputs["valuation"]
    base = _dcf(latest, SCENARIOS["Base"])
    peer_count = len([p for p in inputs["peers"].get("peers", []) if p.get("price_to_sales") or p.get("ev_to_ebitda")])
    holder_count = len(inputs["holders"].get("holders", []))
    annual_rows = len(inputs["annual"].get("periods", []))
    quarterly_rows = len(inputs["quarterly"].get("periods", []))
    wb = Workbook()
    ws = wb.active
    ws.title = "Challenger"
    ws.append(["Area", "Check", "Status", "Evidence", "Suggested repair"])
    checks = [
        ("Statements", "annual_rows_at_least_3", "pass" if annual_rows >= 3 else "fail", annual_rows, "Rerun statement parser"),
        ("Statements", "quarterly_rows_at_least_2", "pass" if quarterly_rows >= 2 else "fail", quarterly_rows, "Rerun statement parser / parse raw 10-Q"),
        ("Market", "market_cap_medium_confidence", "pass" if valuation.get("market_cap_confidence") not in {"low", None, "missing"} else "warning", valuation.get("market_cap_confidence"), "Try alternate market data source"),
        ("Comps", "peer_multiples_available", "pass" if peer_count >= 1 else "fail", peer_count, "Fetch peer multiples"),
        ("Capital signals", "institutional_holder_rows", "pass" if holder_count >= 10 else "warning", holder_count, "Try Nasdaq/13F fallback or paid API"),
        ("DCF", "terminal_value_share", "warning" if base.get("terminal_value_share") and (base["terminal_value_share"] > 0.9 or base["terminal_value_share"] < 0.4) else "pass", base.get("terminal_value_share"), "Review WACC, growth, margin assumptions"),
        (
            "DCF",
            "ev_vs_current_ev_gap",
            "warning",
            f"DCF EV={base.get('enterprise_value')}; current EV={valuation.get('enterprise_value')}",
            "Large gaps should be explained, not hidden",
        ),
    ]
    for row in checks:
        ws.append(row)
    for row in range(2, ws.max_row + 1):
        status = ws.cell(row, 3).value
        if status == "pass":
            _fill_row(ws, row, GREEN)
        elif status == "warning":
            _fill_row(ws, row, YELLOW)
        else:
            _fill_row(ws, row, RED)
    ws = wb.create_sheet("Sample Comparison")
    ws.append(["Sample", "Best use", "Strength", "Weakness"])
    rows = [
        ("Compact DCF + Comps", "Fast initial review", "Small and readable", "Less scenario depth"),
        ("Scenario Architect", "Fuller valuation workflow", "Assumption register and scenarios", "More tabs"),
        ("Comps Focus", "Relative valuation", "Peer stats and implied ranges", "DCF is intentionally secondary"),
        ("Challenger Pack", "Audit companion", "Shows gaps and warnings", "Not a valuation model by itself"),
    ]
    for row in rows:
        ws.append(row)
    ws = wb.create_sheet("Sources")
    _write_sources(ws, target)
    _apply_base_style(wb)
    path = out_dir / f"{target.ticker.lower()}_challenger_pack.xlsx"
    wb.save(path)
    return path


def validate_workbook(path: Path) -> dict[str, Any]:
    wb_formula = load_workbook(path, data_only=False)
    wb_values = load_workbook(path, data_only=True)
    errors: list[str] = []
    visible_numeric_cells = 0
    for ws in wb_formula.worksheets:
        if ws.max_row <= 1 or ws.max_column <= 1:
            errors.append(f"{ws.title}: sparse sheet")
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(token in cell.value for token in ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"]):
                    errors.append(f"{path.name}:{ws.title}!{cell.coordinate}:{cell.value}")
    for ws in wb_values.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    visible_numeric_cells += 1
    return {"path": str(path), "status": "pass" if not errors and visible_numeric_cells > 0 else "needs_review", "errors": errors, "visible_numeric_cells": visible_numeric_cells, "sheets": wb_formula.sheetnames}


def build_samples(target: PublicCompanyTarget, as_of: str | None = None) -> dict[str, Any]:
    date = as_of or TODAY
    inputs = _load_inputs(target)
    out_dir = target.public_company_root / "models" / date / "samples"
    out_dir.mkdir(parents=True, exist_ok=True)
    paths = [
        build_compact_model(target, inputs, out_dir, date),
        build_scenario_model(target, inputs, out_dir, date),
        build_comps_model(target, inputs, out_dir, date),
        build_challenger_pack(target, inputs, out_dir, date),
    ]
    validations = [validate_workbook(path) for path in paths]
    result = {
        "company": target.company_name,
        "ticker": target.ticker,
        "as_of": date,
        "status": "pass" if all(item["status"] == "pass" for item in validations) else "needs_review",
        "samples": [{"name": path.stem, "path": str(path)} for path in paths],
        "validations": validations,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    write_json(out_dir / "model_samples_validation.json", result)
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("target", help="Target key, ticker, or slug")
    parser.add_argument("--as-of", default=None)
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()
    target = PublicCompanyTarget.from_config(load_target(args.target))
    result = build_samples(target, as_of=args.as_of)
    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print(f"{result['status']}: {len(result['samples'])} samples")


if __name__ == "__main__":
    main()
