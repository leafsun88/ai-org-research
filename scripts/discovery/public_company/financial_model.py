"""Create a source-mapped public-company financial model workbook."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

import xlsxwriter
from openpyxl import load_workbook

from .common import PublicCompanyTarget, TODAY, read_json, write_json, write_markdown


SHEET_NAMES = [
    "Summary",
    "Historical Statements",
    "Operating Metrics",
    "Forecast",
    "DCF",
    "Comps",
    "Sensitivity",
    "Capital Signals",
    "Source Map",
    "Assumptions",
]

METRICS = [
    "revenue",
    "gross_profit",
    "operating_income",
    "net_income",
    "operating_cash_flow",
    "capital_expenditures",
    "cash",
    "debt",
    "assets",
    "liabilities",
    "stockholders_equity",
    "shares_outstanding",
]

DEFAULT_ASSUMPTIONS = {
    "revenue_growth": {"value": 0.05, "status": "auto_default", "description": "Base forecast revenue growth"},
    "operating_margin": {"value": 0.12, "status": "auto_default", "description": "Base forecast operating margin"},
    "tax_rate": {"value": 0.21, "status": "auto_default", "description": "Cash tax rate"},
    "da_percent_revenue": {"value": 0.04, "status": "auto_default", "description": "D&A as % of revenue"},
    "capex_percent_revenue": {"value": 0.08, "status": "auto_default", "description": "Capital expenditure as % of revenue"},
    "nwc_percent_revenue": {"value": 0.02, "status": "auto_default", "description": "Change in working capital as % of revenue"},
    "wacc": {"value": 0.10, "status": "auto_default", "description": "Weighted average cost of capital"},
    "terminal_growth": {"value": 0.03, "status": "auto_default", "description": "Terminal growth rate"},
}

ANNUAL_REQUIRED_METRICS = [
    "revenue",
    "gross_profit",
    "operating_income",
    "net_income",
    "operating_cash_flow",
    "capital_expenditures",
    "cash",
    "debt",
    "assets",
    "liabilities",
    "stockholders_equity",
    "shares_outstanding",
]


def _load_inputs(target: PublicCompanyTarget) -> dict[str, Any]:
    root = target.public_company_root
    annual = read_json(root / "data" / "statements" / "annual_statements.json", {"periods": [], "source_map": {}}) or {}
    quarterly = read_json(root / "data" / "statements" / "quarterly_statements.json", {"periods": [], "source_map": {}}) or {}
    valuation = read_json(root / "data" / "valuation" / "valuation_snapshot.json", {}) or {}
    peers = read_json(root / "data" / "valuation" / "peer_valuation.json", {"peers": []}) or {"peers": []}
    insider = read_json(root / "data" / "capital_signals" / "insider_forms.json", {"forms": []}) or {"forms": []}
    holders = read_json(root / "data" / "capital_signals" / "institutional_holders.json", {"holders": []}) or {"holders": []}
    return {
        "annual": annual,
        "quarterly": quarterly,
        "valuation": valuation,
        "peers": peers,
        "insider": insider,
        "holders": holders,
    }


def _latest_annual(inputs: dict[str, Any]) -> dict[str, Any]:
    periods = inputs.get("annual", {}).get("periods", [])
    return periods[-1] if periods else {}


def _calculate_dcf_outputs(latest: dict[str, Any], assumptions: dict[str, Any]) -> dict[str, Any]:
    revenue = float(latest.get("revenue") or 0)
    cash = float(latest.get("cash") or 0)
    debt = float(latest.get("debt") or 0)
    shares = float(latest.get("shares_outstanding") or 0)
    growth = float(assumptions["revenue_growth"]["value"])
    operating_margin = float(assumptions["operating_margin"]["value"])
    tax_rate = float(assumptions["tax_rate"]["value"])
    da_percent = float(assumptions["da_percent_revenue"]["value"])
    capex_percent = float(assumptions["capex_percent_revenue"]["value"])
    nwc_percent = float(assumptions["nwc_percent_revenue"]["value"])
    wacc = float(assumptions["wacc"]["value"])
    terminal_growth = float(assumptions["terminal_growth"]["value"])
    forecast = []
    current_revenue = revenue
    for year in range(1, 6):
        current_revenue *= 1 + growth
        operating_income = current_revenue * operating_margin
        nopat = operating_income * (1 - tax_rate)
        da = current_revenue * da_percent
        capex = -current_revenue * capex_percent
        nwc = -current_revenue * nwc_percent
        fcf = nopat + da + capex + nwc
        forecast.append({"year": year, "revenue": current_revenue, "free_cash_flow": fcf})
    if wacc <= terminal_growth or not forecast:
        return {"status": "invalid", "reason": "wacc_must_exceed_terminal_growth"}
    pv_fcf = [row["free_cash_flow"] / ((1 + wacc) ** row["year"]) for row in forecast]
    terminal_fcf = forecast[-1]["free_cash_flow"] * (1 + terminal_growth)
    terminal_value = terminal_fcf / (wacc - terminal_growth)
    pv_terminal_value = terminal_value / ((1 + wacc) ** 5)
    enterprise_value = sum(pv_fcf) + pv_terminal_value
    equity_value = enterprise_value + cash - debt
    implied_price = equity_value / shares if shares else None
    return {
        "status": "calculated",
        "forecast": forecast,
        "pv_fcf": pv_fcf,
        "terminal_value": terminal_value,
        "pv_terminal_value": pv_terminal_value,
        "enterprise_value": enterprise_value,
        "equity_value": equity_value,
        "implied_price": implied_price,
        "assumption_status": "auto_default" if all(item.get("status") == "auto_default" for item in assumptions.values()) else "user_or_mixed",
    }


def _safe_ratio(numerator: Any, denominator: Any) -> float | None:
    if numerator in {None, ""} or denominator in {None, "", 0}:
        return None
    try:
        denominator_f = float(denominator)
        if denominator_f == 0:
            return None
        return float(numerator) / denominator_f
    except (TypeError, ValueError):
        return None


def _annual_income_years_from_companyfacts(target: PublicCompanyTarget) -> list[str]:
    companyfacts = read_json(target.public_company_root / "metadata" / "sec_companyfacts.json", {}) or {}
    facts = companyfacts.get("facts", {}).get("us-gaap", {})
    years: set[str] = set()
    for tag in ["RevenueFromContractWithCustomerExcludingAssessedTax", "SalesRevenueNet", "Revenues"]:
        units = (facts.get(tag) or {}).get("units", {})
        for unit_values in units.values():
            for fact in unit_values:
                form = str(fact.get("form") or "")
                fp = str(fact.get("fp") or "")
                end = str(fact.get("end") or "")
                if form in {"10-K", "10-K/A", "20-F", "20-F/A"} and fp == "FY" and len(end) >= 4:
                    years.add(end[:4])
    return sorted(years)


def _sensitivity_equity_value(latest: dict[str, Any], assumptions: dict[str, Any], wacc: float, terminal_growth: float) -> float | None:
    if wacc <= terminal_growth:
        return None
    adjusted = json.loads(json.dumps(assumptions))
    adjusted["wacc"]["value"] = wacc
    adjusted["terminal_growth"]["value"] = terminal_growth
    outputs = _calculate_dcf_outputs(latest, adjusted)
    return outputs.get("equity_value") if outputs.get("status") == "calculated" else None


def _write_xlsx_model(
    workbook_path: Path,
    target: PublicCompanyTarget,
    date: str,
    inputs: dict[str, Any],
    latest: dict[str, Any],
    valuation: dict[str, Any],
    assumptions: dict[str, Any],
    model_outputs: dict[str, Any],
    forecast_rows: list[dict[str, Any]],
) -> None:
    annual_periods = inputs["annual"].get("periods", [])
    latest_row = len(annual_periods) + 1 if annual_periods else 2

    workbook = xlsxwriter.Workbook(str(workbook_path))
    workbook.set_calc_mode("auto")
    header_fmt = workbook.add_format({"bold": True, "font_color": "white", "bg_color": "#1F4E79", "border": 1})
    input_fmt = workbook.add_format({"bg_color": "#FFF2CC", "border": 1})
    text_fmt = workbook.add_format({"border": 1})
    num_fmt = workbook.add_format({"num_format": "#,##0", "border": 1})
    pct_fmt = workbook.add_format({"num_format": "0.0%", "border": 1})
    money_fmt = workbook.add_format({"num_format": "$#,##0", "border": 1})

    sheets = {name: workbook.add_worksheet(name) for name in SHEET_NAMES}

    def write_header(ws, headers: list[str]) -> None:
        for col, value in enumerate(headers):
            ws.write(0, col, value, header_fmt)

    def write_cell(ws, row: int, col: int, value: Any, fmt=None) -> None:
        if isinstance(value, (int, float)):
            ws.write_number(row, col, value, fmt or num_fmt)
        elif value is None:
            ws.write_blank(row, col, None, fmt or text_fmt)
        else:
            ws.write(row, col, value, fmt or text_fmt)

    def write_formula(ws, row: int, col: int, formula: str, cached: Any, fmt=None) -> None:
        ws.write_formula(row, col, formula, fmt or num_fmt, "" if cached is None else cached)

    def set_default_widths(ws, widths: list[int]) -> None:
        for col, width in enumerate(widths):
            ws.set_column(col, col, width)

    ws = sheets["Assumptions"]
    write_header(ws, ["name", "value", "status", "description"])
    assumption_rows: dict[str, int] = {}
    for row_idx, (name, item) in enumerate(assumptions.items(), start=1):
        assumption_rows[name] = row_idx + 1
        ws.write(row_idx, 0, name, text_fmt)
        write_cell(ws, row_idx, 1, item["value"], input_fmt)
        ws.write(row_idx, 2, item["status"], text_fmt)
        ws.write(row_idx, 3, item["description"], text_fmt)
    set_default_widths(ws, [24, 14, 16, 44])

    ws = sheets["Historical Statements"]
    write_header(ws, ["period", "fiscal_year", "fiscal_period", "form", "filed", *METRICS])
    for row_idx, row in enumerate(annual_periods, start=1):
        values = [row.get("period"), row.get("fiscal_year"), row.get("fiscal_period"), row.get("form"), row.get("filed"), *[row.get(metric) for metric in METRICS]]
        for col_idx, value in enumerate(values):
            write_cell(ws, row_idx, col_idx, value)
    set_default_widths(ws, [12, 12, 14, 10, 14, 16, 16, 18, 16, 20, 20, 16, 14, 16, 16, 22, 22])

    ws = sheets["Operating Metrics"]
    write_header(ws, ["period", "gross_margin", "operating_margin", "net_margin", "fcf_margin"])
    for idx, row in enumerate(annual_periods, start=2):
        out_row = idx - 1
        hist_row = idx
        ws.write(out_row, 0, row.get("period"), text_fmt)
        write_formula(ws, out_row, 1, f'=IFERROR(\'Historical Statements\'!G{hist_row}/\'Historical Statements\'!F{hist_row},"")', _safe_ratio(row.get("gross_profit"), row.get("revenue")), pct_fmt)
        write_formula(ws, out_row, 2, f'=IFERROR(\'Historical Statements\'!H{hist_row}/\'Historical Statements\'!F{hist_row},"")', _safe_ratio(row.get("operating_income"), row.get("revenue")), pct_fmt)
        write_formula(ws, out_row, 3, f'=IFERROR(\'Historical Statements\'!I{hist_row}/\'Historical Statements\'!F{hist_row},"")', _safe_ratio(row.get("net_income"), row.get("revenue")), pct_fmt)
        fcf_value = (row.get("operating_cash_flow") or 0) + (row.get("capital_expenditures") or 0)
        write_formula(ws, out_row, 4, f'=IFERROR((\'Historical Statements\'!J{hist_row}+\'Historical Statements\'!K{hist_row})/\'Historical Statements\'!F{hist_row},"")', _safe_ratio(fcf_value, row.get("revenue")), pct_fmt)
    set_default_widths(ws, [12, 18, 20, 16, 16])

    ws = sheets["Forecast"]
    write_header(ws, ["metric", "latest", "Y1", "Y2", "Y3", "Y4", "Y5"])
    latest_revenue = latest.get("revenue") or 0
    forecast_revenue = [row.get("revenue") for row in forecast_rows]
    forecast_operating_income = [value * assumptions["operating_margin"]["value"] if isinstance(value, (int, float)) else None for value in forecast_revenue]
    forecast_nopat = [value * (1 - assumptions["tax_rate"]["value"]) if isinstance(value, (int, float)) else None for value in forecast_operating_income]
    forecast_da = [value * assumptions["da_percent_revenue"]["value"] if isinstance(value, (int, float)) else None for value in forecast_revenue]
    forecast_capex = [-value * assumptions["capex_percent_revenue"]["value"] if isinstance(value, (int, float)) else None for value in forecast_revenue]
    forecast_nwc = [-value * assumptions["nwc_percent_revenue"]["value"] if isinstance(value, (int, float)) else None for value in forecast_revenue]
    forecast_fcf = [row.get("free_cash_flow") for row in forecast_rows]
    forecast_rows_values = [
        ("Revenue", f"='Historical Statements'!F{latest_row}", latest_revenue, ["=B2*(1+Assumptions!$B$2)", "=C2*(1+Assumptions!$B$2)", "=D2*(1+Assumptions!$B$2)", "=E2*(1+Assumptions!$B$2)", "=F2*(1+Assumptions!$B$2)"], forecast_revenue),
        ("Operating Income", f"='Historical Statements'!H{latest_row}", latest.get("operating_income") or 0, ["=C2*Assumptions!$B$3", "=D2*Assumptions!$B$3", "=E2*Assumptions!$B$3", "=F2*Assumptions!$B$3", "=G2*Assumptions!$B$3"], forecast_operating_income),
        ("NOPAT", "", None, ["=C3*(1-Assumptions!$B$4)", "=D3*(1-Assumptions!$B$4)", "=E3*(1-Assumptions!$B$4)", "=F3*(1-Assumptions!$B$4)", "=G3*(1-Assumptions!$B$4)"], forecast_nopat),
        ("D&A", "", None, ["=C2*Assumptions!$B$5", "=D2*Assumptions!$B$5", "=E2*Assumptions!$B$5", "=F2*Assumptions!$B$5", "=G2*Assumptions!$B$5"], forecast_da),
        ("Capex", f"='Historical Statements'!K{latest_row}", latest.get("capital_expenditures") or 0, ["=-C2*Assumptions!$B$6", "=-D2*Assumptions!$B$6", "=-E2*Assumptions!$B$6", "=-F2*Assumptions!$B$6", "=-G2*Assumptions!$B$6"], forecast_capex),
        ("Change in NWC", "", None, ["=-C2*Assumptions!$B$7", "=-D2*Assumptions!$B$7", "=-E2*Assumptions!$B$7", "=-F2*Assumptions!$B$7", "=-G2*Assumptions!$B$7"], forecast_nwc),
        ("Free Cash Flow", f"='Historical Statements'!J{latest_row}+'Historical Statements'!K{latest_row}", (latest.get("operating_cash_flow", 0) + latest.get("capital_expenditures", 0)) if latest else 0, ["=C4+C5+C6+C7", "=D4+D5+D6+D7", "=E4+E5+E6+E7", "=F4+F5+F6+F7", "=G4+G5+G6+G7"], forecast_fcf),
    ]
    for row_idx, (label, latest_formula, latest_cached, formulas, cached_values) in enumerate(forecast_rows_values, start=1):
        ws.write(row_idx, 0, label, text_fmt)
        if latest_formula:
            write_formula(ws, row_idx, 1, latest_formula, latest_cached, money_fmt)
        else:
            write_cell(ws, row_idx, 1, latest_cached, money_fmt)
        for offset, formula in enumerate(formulas, start=2):
            cached = cached_values[offset - 2] if offset - 2 < len(cached_values) else None
            write_formula(ws, row_idx, offset, formula, cached, money_fmt)
    set_default_widths(ws, [20, 16, 16, 16, 16, 16, 16])

    ws = sheets["DCF"]
    write_header(ws, ["Item", "Value"])
    dcf_rows = [
        ("WACC", "=Assumptions!$B$8", assumptions["wacc"]["value"], pct_fmt),
        ("Terminal growth", "=Assumptions!$B$9", assumptions["terminal_growth"]["value"], pct_fmt),
    ]
    for idx, value in enumerate(model_outputs.get("pv_fcf", []), start=1):
        dcf_rows.append((f"PV Y{idx}", f"=Forecast!{chr(66 + idx)}8/(1+$B$2)^{idx}", value, money_fmt))
    terminal_fcf = forecast_rows[-1]["free_cash_flow"] * (1 + assumptions["terminal_growth"]["value"]) if forecast_rows else None
    dcf_rows.extend(
        [
            ("Terminal FCF", "=Forecast!G8*(1+$B$3)", terminal_fcf, money_fmt),
            ("Terminal value", "=$B$9/($B$2-$B$3)", model_outputs.get("terminal_value"), money_fmt),
            ("PV terminal value", "=$B$10/(1+$B$2)^5", model_outputs.get("pv_terminal_value"), money_fmt),
            ("Enterprise value", "=SUM($B$4:$B$8)+$B$11", model_outputs.get("enterprise_value"), money_fmt),
            ("Cash", f"='Historical Statements'!L{latest_row}", latest.get("cash") or 0, money_fmt),
            ("Debt", f"='Historical Statements'!M{latest_row}", latest.get("debt") or 0, money_fmt),
            ("Equity value", "=$B$12+$B$13-$B$14", model_outputs.get("equity_value"), money_fmt),
            ("Shares outstanding", f"='Historical Statements'!Q{latest_row}", latest.get("shares_outstanding") or "", num_fmt),
            ("Implied price", "=IFERROR($B$15/$B$16,\"\")", model_outputs.get("implied_price"), money_fmt),
        ]
    )
    for row_idx, (label, formula, cached, fmt) in enumerate(dcf_rows, start=1):
        ws.write(row_idx, 0, label, text_fmt)
        write_formula(ws, row_idx, 1, formula, cached, fmt)
    set_default_widths(ws, [24, 20])

    ws = sheets["Summary"]
    write_header(ws, ["Item", "Value", "Source"])
    summary_rows = [
        ("Company", target.company_name, "target config", None),
        ("Ticker", target.ticker, "target config", None),
        ("As of", date, "script argument", None),
        ("Market cap", valuation.get("market_cap"), "data/valuation/valuation_snapshot.json", None),
        ("Enterprise value", valuation.get("enterprise_value"), "data/valuation/valuation_snapshot.json", None),
        ("Latest annual revenue", latest.get("revenue"), "data/statements/annual_statements.json", f"='Historical Statements'!F{latest_row}"),
        ("Latest annual net income", latest.get("net_income"), "data/statements/annual_statements.json", f"='Historical Statements'!I{latest_row}"),
        ("DCF implied enterprise value", model_outputs.get("enterprise_value"), "DCF sheet", "=DCF!B12"),
        ("DCF implied equity value", model_outputs.get("equity_value"), "DCF sheet", "=DCF!B15"),
        ("DCF implied price", model_outputs.get("implied_price"), "DCF sheet", "=DCF!B17"),
    ]
    for row_idx, (label, value, source, formula) in enumerate(summary_rows, start=1):
        ws.write(row_idx, 0, label, text_fmt)
        if formula:
            write_formula(ws, row_idx, 1, formula, value, money_fmt)
        else:
            write_cell(ws, row_idx, 1, value, money_fmt if isinstance(value, (int, float)) else text_fmt)
        ws.write(row_idx, 2, source, text_fmt)
    set_default_widths(ws, [28, 20, 44])

    ws = sheets["Comps"]
    write_header(ws, ["ticker", "market_cap", "enterprise_value", "trailing_pe", "price_to_sales", "ev_to_ebitda", "source"])
    for row_idx, peer in enumerate(inputs.get("peers", {}).get("peers", []), start=1):
        values = [peer.get("ticker"), peer.get("market_cap"), peer.get("enterprise_value"), peer.get("trailing_pe"), peer.get("price_to_sales"), peer.get("ev_to_ebitda"), peer.get("source")]
        for col_idx, value in enumerate(values):
            write_cell(ws, row_idx, col_idx, value)
    if not target.peer_tickers:
        ws.write_row(1, 0, ["no peer_tickers configured", "", "", "", "", "", "target config"], text_fmt)
    set_default_widths(ws, [12, 18, 18, 14, 16, 16, 44])

    ws = sheets["Sensitivity"]
    wacc_values = [0.08, 0.09, 0.10, 0.11, 0.12]
    growth_values = [0.01, 0.02, 0.03, 0.04, 0.05]
    write_header(ws, ["Terminal growth \\ WACC", *wacc_values])
    for row_idx, growth in enumerate(growth_values, start=1):
        ws.write_number(row_idx, 0, growth, pct_fmt)
        for col_idx, wacc in enumerate(wacc_values, start=1):
            cached = _sensitivity_equity_value(latest, assumptions, wacc, growth)
            formula = f'=IF({chr(65 + col_idx)}$1>$A{row_idx + 1},SUM(Forecast!C8/(1+{chr(65 + col_idx)}$1)^1,Forecast!D8/(1+{chr(65 + col_idx)}$1)^2,Forecast!E8/(1+{chr(65 + col_idx)}$1)^3,Forecast!F8/(1+{chr(65 + col_idx)}$1)^4,Forecast!G8/(1+{chr(65 + col_idx)}$1)^5)+Forecast!G8*(1+$A{row_idx + 1})/({chr(65 + col_idx)}$1-$A{row_idx + 1})/(1+{chr(65 + col_idx)}$1)^5+DCF!$B$13-DCF!$B$14,"")'
            write_formula(ws, row_idx, col_idx, formula, cached, money_fmt)
    set_default_widths(ws, [24, 16, 16, 16, 16, 16])

    ws = sheets["Capital Signals"]
    write_header(ws, ["type", "name_or_form", "date", "value", "source"])
    row_idx = 1
    for row in inputs.get("insider", {}).get("forms", []):
        for col_idx, value in enumerate(["insider_form", row.get("form"), row.get("filing_date"), row.get("description"), "data/capital_signals/insider_forms.json"]):
            write_cell(ws, row_idx, col_idx, value)
        row_idx += 1
    for row in inputs.get("holders", {}).get("holders", []):
        for col_idx, value in enumerate(["institutional_holder", row.get("Holder") or row.get("holder"), row.get("Date Reported") or row.get("date_reported"), row.get("Shares") or row.get("shares"), "data/capital_signals/institutional_holders.json"]):
            write_cell(ws, row_idx, col_idx, value)
        row_idx += 1
    set_default_widths(ws, [18, 34, 16, 28, 44])

    ws = sheets["Source Map"]
    write_header(ws, ["metric", "source", "status"])
    source_map = inputs.get("annual", {}).get("source_map", {})
    for row_idx, metric in enumerate(METRICS, start=1):
        source = source_map.get(metric) or "data/statements/annual_statements.json"
        status = "present" if latest.get(metric) not in {None, ""} else "missing"
        ws.write(row_idx, 0, metric, text_fmt)
        ws.write(row_idx, 1, json.dumps(source, ensure_ascii=False)[:500], text_fmt)
        ws.write(row_idx, 2, status, text_fmt)
    tail_row = len(METRICS) + 1
    ws.write(tail_row, 0, "market_data", text_fmt)
    ws.write(tail_row, 1, "data/market_data/price_history.csv", text_fmt)
    ws.write(tail_row, 2, "present" if (target.public_company_root / "data" / "market_data" / "price_history.csv").exists() else "missing", text_fmt)
    ws.write(tail_row + 1, 0, "valuation", text_fmt)
    ws.write(tail_row + 1, 1, "data/valuation/valuation_snapshot.json", text_fmt)
    ws.write(tail_row + 1, 2, "present" if valuation else "missing", text_fmt)
    set_default_widths(ws, [24, 70, 14])

    workbook.close()


def write_financial_model(target: PublicCompanyTarget, as_of_date: str | None = None) -> dict[str, str]:
    date = as_of_date or TODAY
    root = target.public_company_root
    model_dir = root / "models" / date
    model_dir.mkdir(parents=True, exist_ok=True)
    inputs = _load_inputs(target)
    annual_periods = inputs["annual"].get("periods", [])
    latest = _latest_annual(inputs)
    valuation = inputs["valuation"]
    assumptions = DEFAULT_ASSUMPTIONS
    model_outputs = _calculate_dcf_outputs(latest, assumptions)
    forecast_rows = model_outputs.get("forecast", []) if model_outputs.get("status") == "calculated" else []

    workbook_path = model_dir / "financial_model.xlsx"
    _write_xlsx_model(workbook_path, target, date, inputs, latest, valuation, assumptions, model_outputs, forecast_rows)

    model_inputs = {
        "company": target.company_name,
        "ticker": target.ticker,
        "as_of_date": date,
        "assumptions": assumptions,
        "latest_annual": latest,
        "valuation": valuation,
        "model_outputs": model_outputs,
        "peer_tickers": target.peer_tickers,
        "source_files": {
            "annual": "data/statements/annual_statements.json",
            "quarterly": "data/statements/quarterly_statements.json",
            "market_data": "data/market_data/price_history.csv",
            "valuation": "data/valuation/valuation_snapshot.json",
        },
    }
    inputs_path = write_json(model_dir / "model_inputs.json", model_inputs)
    validation = validate_model(target, workbook_path, model_dir)
    validation_path = write_json(model_dir / "model_validation.json", validation)
    gap_lines = [f"- {gap}" for gap in validation.get("gaps", [])] or ["- none"]
    warning_lines = [f"- {warning}" for warning in validation.get("warnings", [])] or ["- none"]
    write_json(root / "metadata" / "model_quality_audit.json", validation)
    write_markdown(
        root / "metadata" / "model_quality_audit.md",
        [
            f"# {target.company_name} Model Quality Audit",
            "",
            f"- Status: {validation['status']}",
            f"- Model usability: {validation.get('model_usability', 'unknown')}",
            "",
            "## Gaps",
            *gap_lines,
            "",
            "## Warnings",
            *warning_lines,
        ],
    )
    summary_path = write_markdown(
        model_dir / "model_summary.md",
        [
            f"# {target.company_name} Financial Model",
            "",
            f"- Workbook: `{workbook_path.name}`",
            f"- As of: {date}",
            f"- Latest annual revenue: {latest.get('revenue', 'missing')}",
            f"- Market cap: {valuation.get('market_cap', 'missing')}",
            f"- Validation status: {validation['status']}",
            "",
            "## Gaps",
            *gap_lines,
            "",
            "## Warnings",
            *warning_lines,
        ],
    )
    return {
        "workbook_path": str(workbook_path),
        "inputs_path": str(inputs_path),
        "summary_path": str(summary_path),
        "validation_path": str(validation_path),
    }


def validate_model(target: PublicCompanyTarget, workbook_path: Path, model_dir: Path) -> dict[str, Any]:
    root = target.public_company_root
    checks: dict[str, dict[str, Any]] = {}
    gaps: list[str] = []
    warnings: list[str] = []
    fatal_errors: list[str] = []
    checks["workbook_created"] = {"passed": workbook_path.exists(), "path": str(workbook_path)}
    try:
        wb = load_workbook(workbook_path, data_only=False)
        checks["required_sheets"] = {"passed": wb.sheetnames == SHEET_NAMES, "sheets": wb.sheetnames}
        formula_errors = []
        formula_count = 0
        sheet_formula_counts: dict[str, int] = {}
        for ws in wb.worksheets:
            sheet_formula_counts[ws.title] = 0
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formula_count += 1
                        sheet_formula_counts[ws.title] += 1
                    if isinstance(value, str) and any(err in value for err in ["#REF!", "#DIV/0!", "#VALUE!"]):
                        formula_errors.append(f"{ws.title}!{cell.coordinate}: {value}")
        checks["formula_error_tokens"] = {"passed": not formula_errors, "errors": formula_errors[:20]}
        if formula_errors:
            fatal_errors.extend(formula_errors[:20])
        formula_sheets = ["Summary", "Forecast", "DCF", "Sensitivity"]
        formula_sheets_with_links = [sheet for sheet in formula_sheets if sheet_formula_counts.get(sheet, 0) > 0]
        checks["workbook_formula_links"] = {
            "passed": formula_count >= 10 and len(formula_sheets_with_links) >= 3,
            "formula_count": formula_count,
            "sheet_formula_counts": sheet_formula_counts,
            "required_formula_sheets": formula_sheets,
            "formula_sheets_with_links": formula_sheets_with_links,
        }
        wb_values = load_workbook(workbook_path, data_only=True)
        metric_ws = wb_values["Operating Metrics"]
        visible_metric_values = 0
        metric_rows = 0
        for row in metric_ws.iter_rows(min_row=2, values_only=True):
            if row[0] in {None, ""}:
                continue
            metric_rows += 1
            visible_metric_values += sum(1 for value in row[1:5] if isinstance(value, (int, float)))
        checks["operating_metrics_visible_values"] = {
            "passed": metric_rows == 0 or visible_metric_values >= metric_rows,
            "metric_rows": metric_rows,
            "visible_numeric_values": visible_metric_values,
            "minimum_visible_numeric_values": metric_rows,
        }
    except Exception as exc:  # noqa: BLE001
        checks["workbook_openable"] = {"passed": False, "error": str(exc)}
        fatal_errors.append(str(exc))
    else:
        checks["workbook_openable"] = {"passed": True}

    annual = read_json(root / "data" / "statements" / "annual_statements.json", {"periods": []}) or {"periods": []}
    quarterly = read_json(root / "data" / "statements" / "quarterly_statements.json", {"periods": []}) or {"periods": []}
    valuation = read_json(root / "data" / "valuation" / "valuation_snapshot.json", {}) or {}
    peers = read_json(root / "data" / "valuation" / "peer_valuation.json", {"peers": []}) or {"peers": []}
    latest = (annual.get("periods") or [{}])[-1] if annual.get("periods") else {}
    checks["annual_statements"] = {"passed": bool(annual.get("periods")), "periods": len(annual.get("periods", []))}
    checks["quarterly_statements"] = {"passed": bool(quarterly.get("periods")), "periods": len(quarterly.get("periods", []))}
    checks["market_data"] = {"passed": (root / "data" / "market_data" / "price_history.csv").exists()}
    checks["valuation"] = {"passed": (root / "data" / "valuation" / "valuation_snapshot.json").exists()}
    missing_annual = [metric for metric in ANNUAL_REQUIRED_METRICS if latest.get(metric) in {None, ""}]
    checks["annual_metric_completeness"] = {
        "passed": not missing_annual,
        "missing": missing_annual,
        "required": ANNUAL_REQUIRED_METRICS,
    }
    expected_annual_years = _annual_income_years_from_companyfacts(target)
    actual_annual_income_years = sorted(
        {
            str(row.get("period"))
            for row in annual.get("periods", [])
            if row.get("revenue") not in {None, ""} and (row.get("operating_income") not in {None, ""} or row.get("net_income") not in {None, ""})
        }
    )
    expected_history_count = min(3, len(expected_annual_years)) if expected_annual_years else min(3, len(actual_annual_income_years))
    checks["annual_history_coverage"] = {
        "passed": len(actual_annual_income_years) >= expected_history_count,
        "expected_years_from_companyfacts": expected_annual_years,
        "actual_income_years": actual_annual_income_years,
        "minimum_required": expected_history_count,
    }
    quarterly_income_rows = [
        row
        for row in quarterly.get("periods", [])
        if row.get("revenue") not in {None, ""} and (row.get("net_income") not in {None, ""} or row.get("operating_income") not in {None, ""})
    ]
    checks["quarterly_income_statement"] = {
        "passed": bool(quarterly_income_rows),
        "rows_with_revenue_and_income": len(quarterly_income_rows),
        "periods": len(quarterly.get("periods", [])),
    }
    market_cap_source = str(valuation.get("market_cap_source") or "")
    market_cap_confidence = str(valuation.get("market_cap_confidence") or "unknown")
    checks["market_cap_confidence"] = {
        "passed": market_cap_confidence not in {"low", "missing", "unknown"} and "weighted_average" not in market_cap_source,
        "source": market_cap_source or valuation.get("source"),
        "confidence": market_cap_confidence,
    }
    peer_rows = peers.get("peers", [])
    peer_rows_with_multiples = [
        row
        for row in peer_rows
        if row.get("price_to_sales") not in {None, ""} or row.get("ev_to_ebitda") not in {None, ""} or row.get("trailing_pe") not in {None, ""}
    ]
    checks["comps_multiples"] = {
        "passed": bool(peer_rows_with_multiples),
        "peers": len(peer_rows),
        "peers_with_multiples": len(peer_rows_with_multiples),
    }
    model_inputs = read_json(model_dir / "model_inputs.json", {}) or {}
    model_outputs = model_inputs.get("model_outputs", {})
    checks["deterministic_dcf_outputs"] = {
        "passed": model_outputs.get("status") == "calculated" and model_outputs.get("implied_price") is not None,
        "status": model_outputs.get("status"),
        "implied_price": model_outputs.get("implied_price"),
    }
    if model_outputs.get("assumption_status") == "auto_default":
        warnings.append("dcf_uses_auto_default_assumptions")
    if not checks.get("workbook_formula_links", {}).get("passed"):
        warnings.append("workbook_static_values_only_not_a_full_formula_model")
    for name, check in checks.items():
        if not check.get("passed"):
            gaps.append(name)
    status = "pass" if not fatal_errors and not gaps else "needs_review"
    if "workbook_formula_links" in gaps:
        model_usability = "static_analysis_pack"
    else:
        model_usability = "decision_ready" if status == "pass" and not warnings else "draft_model" if not fatal_errors and len(gaps) <= 2 else "workbook_shell_only"
    return {
        "company": target.company_name,
        "ticker": target.ticker,
        "status": status,
        "model_usability": model_usability,
        "checks": checks,
        "gaps": gaps,
        "warnings": warnings,
        "fatal_errors": fatal_errors,
        "next_steps": [
            "Review auto_default assumptions before relying on valuation output.",
            "Add paid transcript or ASR transcript source if official IR transcript remains unavailable.",
            "Add explicit peer_tickers in config before using comps output for decisions.",
        ],
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }


def main(argv: list[str] | None = None) -> int:
    import argparse
    import sys

    from target_config import load_target

    parser = argparse.ArgumentParser()
    parser.add_argument("target", help="Configured research key, ticker, or slug")
    parser.add_argument("--as-of-date", default=None)
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args(argv)

    config = load_target(args.target)
    target = PublicCompanyTarget.from_config(config)
    result = write_financial_model(target, as_of_date=args.as_of_date)
    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print(result["workbook_path"])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
